from datetime import datetime
from openpyxl import load_workbook

from src.models import Scheme, SchemeCache


def load_schemes(scheme_list):
    wb = load_workbook(filename=scheme_list, read_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=7, min_col=1, max_row=1000000, max_col=7):
        row = [ele.value for ele in row]
        if "<EmptyCell>" in str(list(row)[0]):
            break
        scheme = get_scheme_from_row(row)
        SchemeCache().add(scheme)
    wb.close()


def get_scheme_from_row(row):
    scheme = Scheme()
    scheme.name = row[0]
    scheme.category = row[1]
    scheme.latest_nav_date = datetime.strptime(row[2], "%d-%m-%Y")
    scheme.latest_nav = float(row[3])
    scheme.previous_nav_date = datetime.strptime(row[4], "%d-%m-%Y")
    scheme.previous_nav = float(row[5])
    scheme.change_percentage = float(row[6])
    return scheme
